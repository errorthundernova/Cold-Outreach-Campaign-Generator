"""
Generator for the Excel outreach tracker file.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter
import datetime
from config.icp_config import PAIN_POINTS

def generate_excel_tracker(prospects: list[dict], output_path: str):
    """
    Generates an Excel tracker with prospects data and analysis.
    
    Args:
        prospects (list[dict]): List of prospect dictionaries.
        output_path (str): File path to save the Excel file.
    """
    wb = Workbook()
    
    # --- Sheet 1: Outreach Tracker ---
    ws1 = wb.active
    ws1.title = "Outreach Tracker"
    
    headers = [
        "Lead_ID", "Name", "Title", "Company", "Industry", "Company_Size", "Cloud_Stack",
        "Est_Monthly_Spend", "Lead_Source", "Pain_Point", "Email_Step", "LinkedIn_Step",
        "Call_Attempted", "Status", "Priority_Tier", "Next_Action_Date", "Response_Notes", "Demo_Booked"
    ]
    
    ws1.append(headers)
    
    # Header styling
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for cell in ws1[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        
    ws1.freeze_panes = "A2"
    
    # Data Validation
    dv_status = DataValidation(type="list", formula1='"Cold,Contacted,Follow-Up,Demo Booked,No Response,Disqualified"', allow_blank=True)
    dv_priority = DataValidation(type="list", formula1='"High,Medium,Low"', allow_blank=True)
    ws1.add_data_validation(dv_status)
    ws1.add_data_validation(dv_priority)
    
    today_plus_one = (datetime.date.today() + datetime.timedelta(days=1)).isoformat()
    
    row_idx = 2
    for prospect in prospects:
        row = [
            prospect.get("lead_id"),
            prospect.get("name"),
            prospect.get("title"),
            prospect.get("company"),
            prospect.get("industry"),
            prospect.get("company_size"),
            prospect.get("cloud_stack"),
            prospect.get("estimated_monthly_spend"),
            prospect.get("lead_source"),
            prospect.get("pain_point"),
            0,                      # Email_Step
            "Not sent",             # LinkedIn_Step
            False,                  # Call_Attempted
            "Cold",                 # Status
            prospect.get("priority_tier", "Low"), # Priority_Tier
            today_plus_one,         # Next_Action_Date
            "",                     # Response_Notes
            False                   # Demo_Booked
        ]
        ws1.append(row)
        
        # Add Data Validation to specific cells
        dv_status.add(f"N{row_idx}")
        dv_priority.add(f"O{row_idx}")
        
        # Alternating row colors
        if row_idx % 2 != 0:
            row_fill = PatternFill(start_color="F9F9F9", end_color="F9F9F9", fill_type="solid")
            for col in range(1, len(headers) + 1):
                ws1.cell(row=row_idx, column=col).fill = row_fill
                
        row_idx += 1

    # Autosize columns
    for col in ws1.columns:
        max_length = 0
        column_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = max(12, min(max_length + 2, 40))
        ws1.column_dimensions[column_letter].width = adjusted_width

    # Conditional Formatting for Status
    status_col = "N"
    blue_fill = PatternFill(start_color="DDEEFF", end_color="DDEEFF", fill_type="solid")
    orange_fill = PatternFill(start_color="FFE8CC", end_color="FFE8CC", fill_type="solid")
    green_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
    gray_fill = PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")
    bold_font = Font(bold=True)
    
    ws1.conditional_formatting.add(f'{status_col}2:{status_col}1000', CellIsRule(operator='equal', formula=['"Contacted"'], fill=blue_fill))
    ws1.conditional_formatting.add(f'{status_col}2:{status_col}1000', CellIsRule(operator='equal', formula=['"Follow-Up"'], fill=orange_fill))
    ws1.conditional_formatting.add(f'{status_col}2:{status_col}1000', CellIsRule(operator='equal', formula=['"No Response"'], fill=gray_fill))
    ws1.conditional_formatting.add(f'{status_col}2:{status_col}1000', CellIsRule(operator='equal', formula=['"Demo Booked"'], fill=green_fill, font=bold_font))

    # Conditional Formatting for Priority
    priority_col = "O"
    high_fill = PatternFill(start_color="FFB3B3", end_color="FFB3B3", fill_type="solid")
    high_font = Font(color="8B0000")
    med_fill = PatternFill(start_color="FFF2B3", end_color="FFF2B3", fill_type="solid")
    med_font = Font(color="7B6000")
    low_fill = PatternFill(start_color="B3FFB3", end_color="B3FFB3", fill_type="solid")
    low_font = Font(color="1B5E20")
    
    ws1.conditional_formatting.add(f'{priority_col}2:{priority_col}1000', CellIsRule(operator='equal', formula=['"High"'], fill=high_fill, font=high_font))
    ws1.conditional_formatting.add(f'{priority_col}2:{priority_col}1000', CellIsRule(operator='equal', formula=['"Medium"'], fill=med_fill, font=med_font))
    ws1.conditional_formatting.add(f'{priority_col}2:{priority_col}1000', CellIsRule(operator='equal', formula=['"Low"'], fill=low_fill, font=low_font))

    # --- Sheet 2: Pain Point Mapping ---
    ws2 = wb.create_sheet(title="Pain Point Mapping")
    ws2_headers = [
        "Lead_ID", "Name", "Title", "Company", "Cloud_Stack", "Est_Monthly_Spend",
        "Primary_Pain_Point", "Secondary_Pain_Point", "Mapped_Value_Prop", "Outreach_Angle"
    ]
    ws2.append(ws2_headers)
    
    for cell in ws2[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        
    ws2.freeze_panes = "A2"
    
    for prospect in prospects:
        cloud_stack = prospect.get("cloud_stack", "AWS")
        pain_points_list = PAIN_POINTS.get(cloud_stack, ["Unoptimized infrastructure", "Wasteful spending"])
        primary = pain_points_list[0] if len(pain_points_list) > 0 else "High costs"
        secondary = pain_points_list[1] if len(pain_points_list) > 1 else "Visibility issues"
        
        mapped_value_prop = f"CloudKeeper directly resolves {primary.lower()} automatically."
        angle = "Cost Savings"
        if "visibility" in primary.lower():
            angle = "Visibility"
        elif "compliance" in primary.lower():
            angle = "Compliance"
            
        ws2.append([
            prospect.get("lead_id"),
            prospect.get("name"),
            prospect.get("title"),
            prospect.get("company"),
            cloud_stack,
            prospect.get("estimated_monthly_spend"),
            primary,
            secondary,
            mapped_value_prop,
            angle
        ])
    
    for col in ws2.columns:
        max_length = 0
        column_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = max(12, min(max_length + 2, 40))
        ws2.column_dimensions[column_letter].width = adjusted_width

    # --- Sheet 3: Performance Analysis ---
    ws3 = wb.create_sheet(title="Performance Analysis")
    
    section_fill = PatternFill(start_color="2E4057", end_color="2E4057", fill_type="solid")
    
    def add_section_header(row, title):
        cell = ws3.cell(row=row, column=1, value=title)
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = section_fill
        ws3.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        return row + 1

    current_row = 1
    current_row = add_section_header(current_row, "Section A — Pipeline Summary")
    
    ws3.cell(row=current_row, column=1, value="Total Prospects:")
    ws3.cell(row=current_row, column=2, value=len(prospects))
    current_row += 2
    
    ws3.cell(row=current_row, column=1, value="Status").font = Font(bold=True)
    ws3.cell(row=current_row, column=2, value="Count").font = Font(bold=True)
    ws3.cell(row=current_row, column=3, value="%").font = Font(bold=True)
    current_row += 1
    
    statuses = ["Cold", "Contacted", "Follow-Up", "Demo Booked", "No Response", "Disqualified"]
    for status in statuses:
        ws3.cell(row=current_row, column=1, value=status)
        ws3.cell(row=current_row, column=2, value=f'=COUNTIF(\'Outreach Tracker\'!N:N, "{status}")')
        ws3.cell(row=current_row, column=3, value=f'=B{current_row}/{max(1, len(prospects))}')
        ws3.cell(row=current_row, column=3).number_format = '0.00%'
        current_row += 1
    
    current_row += 1
    current_row = add_section_header(current_row, "Section B — Source Analysis")
    ws3.cell(row=current_row, column=1, value="Lead Source").font = Font(bold=True)
    ws3.cell(row=current_row, column=2, value="Count").font = Font(bold=True)
    ws3.cell(row=current_row, column=3, value="% of total").font = Font(bold=True)
    current_row += 1
    
    sources = list(set([p.get("lead_source") for p in prospects if p.get("lead_source")]))
    for src in sources:
        count = sum(1 for p in prospects if p.get("lead_source") == src)
        ws3.cell(row=current_row, column=1, value=src)
        ws3.cell(row=current_row, column=2, value=count)
        ws3.cell(row=current_row, column=3, value=count/max(1, len(prospects))).number_format = '0.00%'
        current_row += 1
    
    current_row += 1
    current_row = add_section_header(current_row, "Section C — Industry and Cloud Stack")
    ws3.cell(row=current_row, column=1, value="Industry").font = Font(bold=True)
    ws3.cell(row=current_row, column=2, value="Count").font = Font(bold=True)
    current_row += 1
    
    industries = list(set([p.get("industry") for p in prospects if p.get("industry")]))
    for ind in industries:
        count = sum(1 for p in prospects if p.get("industry") == ind)
        ws3.cell(row=current_row, column=1, value=ind)
        ws3.cell(row=current_row, column=2, value=count)
        current_row += 1
        
    current_row += 1
    ws3.cell(row=current_row, column=1, value="Cloud Stack").font = Font(bold=True)
    ws3.cell(row=current_row, column=2, value="Count").font = Font(bold=True)
    current_row += 1
    
    stacks = list(set([p.get("cloud_stack") for p in prospects if p.get("cloud_stack")]))
    for stack in stacks:
        count = sum(1 for p in prospects if p.get("cloud_stack") == stack)
        ws3.cell(row=current_row, column=1, value=stack)
        ws3.cell(row=current_row, column=2, value=count)
        current_row += 1

    current_row += 1
    current_row = add_section_header(current_row, "Section D — Simulated Funnel")
    ws3.cell(row=current_row, column=1, value="Prospects Contacted")
    ws3.cell(row=current_row, column=2, value=f'=COUNTIFS(\'Outreach Tracker\'!N:N, "<>Cold", \'Outreach Tracker\'!N:N, "<>Status")')
    current_row += 1
    ws3.cell(row=current_row, column=1, value="Replies Received")
    ws3.cell(row=current_row, column=2, value="[Update manually]")
    current_row += 1
    ws3.cell(row=current_row, column=1, value="Demos Booked")
    ws3.cell(row=current_row, column=2, value=f'=COUNTIF(\'Outreach Tracker\'!R:R, TRUE)')
    current_row += 1
    ws3.cell(row=current_row, column=1, value="Overall Conversion")
    ws3.cell(row=current_row, column=2, value=f'=B{current_row-1}/{max(1, len(prospects))}')
    ws3.cell(row=current_row, column=2).number_format = '0.00%'
    
    for col_idx in range(1, ws3.max_column + 1):
        ws3.column_dimensions[get_column_letter(col_idx)].width = 25
        
    wb.save(output_path)
